#!/usr/bin/env python3
"""Import Star Chef v0.8a supplier quote workbooks.

This importer intentionally produces two outputs:

1. A public, sanitized JavaScript summary for the GitHub Pages demo.
2. A private JSON file with raw rows and exact prices for local audit only.

The public summary must not expose full supplier quote rows or exact item-level
prices from the workbook.
"""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from statistics import mean
from typing import Any

from openpyxl import load_workbook


DEFAULT_SOURCE = (
    "/Users/marlins/Library/Containers/com.tencent.WeWorkMac/Data/Documents/Profiles/"
    "F02957C04AEB6A4433AA68D86C18DB18/Caches/Files/2026-06/"
    "4fb157473a98f1bae84608f00f00cf27/蓝鼎2026年6月供应商报价单.xlsx"
)

COLUMN_KEYS = [
    "item_code",
    "item_name",
    "spec",
    "supplier",
    "purchase_unit",
    "price_with_tax_delivery",
    "start_date",
    "end_date",
]

CATEGORY_BY_PREFIX = {
    "A01": "蔬菜菌菇",
    "A02": "米面粮油",
    "A03": "肉禽肉类",
    "A04": "水产海鲜",
    "A05": "水果",
    "A06": "蛋奶豆制品",
    "A07": "预包装食品",
    "A08": "调味干货",
    "A09": "半成品冻品",
    "B01": "饮品乳品",
    "C01": "酒水调料",
}

MENU_KEYWORDS = [
    "毛豆",
    "春笋",
    "笋",
    "小排",
    "仔排",
    "排骨",
    "里脊",
    "鸡",
    "鸭",
    "牛",
    "虾",
    "鱼",
    "土豆",
    "青椒",
    "西红柿",
    "番茄",
    "菠菜",
    "豆芽",
    "山药",
    "藕",
    "茄子",
    "豆腐",
    "鸡蛋",
]


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def normalize_price(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return round(float(value), 4)
    except (TypeError, ValueError):
        return None


def normalize_date(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    if re.fullmatch(r"\d{8}", text):
        return f"{text[:4]}-{text[4:6]}-{text[6:8]}"
    return text


def category_for_code(item_code: str) -> str:
    return CATEGORY_BY_PREFIX.get(item_code[:3], "其他")


def price_band(price: float | None, unit: str) -> str:
    if price is None:
        return "缺价格"
    if price <= 5:
        band = "≤5元"
    elif price <= 10:
        band = "5-10元"
    elif price <= 20:
        band = "10-20元"
    elif price <= 40:
        band = "20-40元"
    else:
        band = ">40元"
    if unit == "均值":
        return f"均值：{band}"
    unit_text = f"/{unit}" if unit else ""
    return f"{band}{unit_text}"


def read_sheet_rows(source: Path, sheet_name: str) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    workbook = load_workbook(source, data_only=True, read_only=True)
    if sheet_name not in workbook.sheetnames:
      raise ValueError(f"Sheet not found: {sheet_name}. Available sheets: {', '.join(workbook.sheetnames)}")

    ws = workbook[sheet_name]
    header = [normalize_text(value) for value in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows: list[dict[str, Any]] = []

    for row_number, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        raw = dict(zip(COLUMN_KEYS, list(values[: len(COLUMN_KEYS)])))
        item_code = normalize_text(raw.get("item_code"))
        item_name = normalize_text(raw.get("item_name"))
        if not item_code and not item_name:
            continue

        purchase_unit = normalize_text(raw.get("purchase_unit"))
        price = normalize_price(raw.get("price_with_tax_delivery"))
        spec = normalize_text(raw.get("spec"))
        supplier = normalize_text(raw.get("supplier"))

        row = {
            "row_number": row_number,
            "item_code": item_code,
            "item_name": item_name,
            "spec": spec,
            "supplier": supplier,
            "purchase_unit": purchase_unit,
            "price_with_tax_delivery": price,
            "start_date": normalize_date(raw.get("start_date")),
            "end_date": normalize_date(raw.get("end_date")),
            "category": category_for_code(item_code),
            "gaps": [],
        }

        if not spec:
            row["gaps"].append("缺规格")
        if not purchase_unit:
            row["gaps"].append("缺采购单位")
        if price is None:
            row["gaps"].append("缺价格")
        if not supplier:
            row["gaps"].append("缺供应商")
        if not row["start_date"]:
            row["gaps"].append("缺开始时间")
        if not row["end_date"]:
            row["gaps"].append("缺结束时间")

        rows.append(row)

    workbook.close()
    meta = {
        "source_file": source.name,
        "sheet_name": sheet_name,
        "header": header[: len(COLUMN_KEYS)],
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    return rows, meta


def candidate_score(row: dict[str, Any]) -> int:
    name = row["item_name"]
    score = 0
    for keyword in MENU_KEYWORDS:
        if keyword in name:
            score += 8
    if row["category"] in {"蔬菜菌菇", "肉禽肉类", "水产海鲜", "蛋奶豆制品", "半成品冻品"}:
        score += 4
    if row["price_with_tax_delivery"] is not None:
        score += 3
    if row["gaps"]:
        score -= len(row["gaps"])
    return score


def mapping_hint(row: dict[str, Any]) -> str:
    category = row["category"]
    if category == "蔬菜菌菇":
        return "优先映射食材库；适合素菜、小荤配色和时令候选。"
    if category == "肉禽肉类":
        return "优先映射菜品 BOM；需要成本卡、出成率和SOP。"
    if category == "水产海鲜":
        return "需要过敏原、留样和价格波动门禁。"
    if category == "半成品冻品":
        return "需要复热SOP、出品克重和试推评分。"
    if category == "米面粮油":
        return "进入基础物料库；不直接生成菜品但影响成本底线。"
    return "进入标准物料库，等待项目规则或菜谱引用。"


def build_summary(rows: list[dict[str, Any]], meta: dict[str, Any]) -> dict[str, Any]:
    total = len(rows)
    price_missing = [row for row in rows if "缺价格" in row["gaps"]]
    spec_missing = [row for row in rows if "缺规格" in row["gaps"]]
    unit_missing = [row for row in rows if "缺采购单位" in row["gaps"]]
    supplier_count = len({row["supplier"] for row in rows if row["supplier"]})
    category_counts = Counter(row["category"] for row in rows)
    prefix_counts = Counter(row["item_code"][:3] if row["item_code"] else "missing" for row in rows)
    date_windows = Counter((row["start_date"], row["end_date"]) for row in rows)

    category_quality = []
    by_category: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        by_category[row["category"]].append(row)

    for category, cat_rows in sorted(by_category.items(), key=lambda item: (-len(item[1]), item[0])):
        priced = [row for row in cat_rows if row["price_with_tax_delivery"] is not None]
        prices = [row["price_with_tax_delivery"] for row in priced]
        category_quality.append({
            "category": category,
            "row_count": len(cat_rows),
            "priced_count": len(priced),
            "missing_price_count": len(cat_rows) - len(priced),
            "price_coverage": round(len(priced) / len(cat_rows), 4) if cat_rows else 0,
            "public_price_band": price_band(mean(prices), "均值") if prices else "缺价格",
        })

    candidate_rows = sorted(rows, key=candidate_score, reverse=True)
    public_candidates = []
    seen_names = set()
    for row in candidate_rows:
        if candidate_score(row) <= 3 or row["item_name"] in seen_names:
            continue
        seen_names.add(row["item_name"])
        public_candidates.append({
            "item_code": row["item_code"],
            "item_name": row["item_name"],
            "category": row["category"],
            "spec_status": "有规格" if row["spec"] else "缺规格",
            "purchase_unit": row["purchase_unit"],
            "price_band": price_band(row["price_with_tax_delivery"], row["purchase_unit"]),
            "quote_window": f"{row['start_date']} 至 {row['end_date']}",
            "mapping_hint": mapping_hint(row),
            "gaps": row["gaps"],
        })
        if len(public_candidates) >= 18:
            break

    gate_status = [
        {
            "gate": "原始行保存",
            "status": "pass",
            "value": f"{total} 行",
            "detail": "本地私有 JSON 已保存完整原始行和精确报价；公开页面只显示脱敏摘要。",
        },
        {
            "gate": "价格完整性",
            "status": "review" if price_missing else "pass",
            "value": f"{len(price_missing)} 行缺价格",
            "detail": "缺价格行不能进入自动菜单候选，需要采购经理补价或标记停用。",
        },
        {
            "gate": "规格完整性",
            "status": "review" if spec_missing else "pass",
            "value": f"{len(spec_missing)} 行缺规格",
            "detail": "缺规格会影响单位换算、成本卡和采购拆分。",
        },
        {
            "gate": "供应商主体",
            "status": "pass" if supplier_count > 0 else "blocked",
            "value": f"{supplier_count} 家",
            "detail": "供应商主体可用于后续保供评分、账期和履约分析。",
        },
    ]

    return {
        "version": "v0.8a-import",
        "privacy_mode": "public_sanitized",
        "source_file": meta["source_file"],
        "sheet_name": meta["sheet_name"],
        "generated_at": meta["generated_at"],
        "column_interpretation": {
            "note": "上海 sheet 第 7 列表头重复写成“结束时间”，导入器按列位置解释为开始时间。",
            "columns": COLUMN_KEYS,
            "source_header": meta["header"],
        },
        "summary": {
            "row_count": total,
            "supplier_count": supplier_count,
            "category_count": len(category_counts),
            "missing_price_count": len(price_missing),
            "missing_spec_count": len(spec_missing),
            "missing_unit_count": len(unit_missing),
            "primary_quote_windows": [
                {"start_date": start, "end_date": end, "row_count": count}
                for (start, end), count in date_windows.most_common(5)
            ],
        },
        "category_counts": [
            {"category": category, "row_count": count}
            for category, count in category_counts.most_common()
        ],
        "prefix_counts": [
            {"prefix": prefix, "row_count": count}
            for prefix, count in prefix_counts.most_common()
        ],
        "category_quality": category_quality,
        "public_candidates": public_candidates,
        "gate_status": gate_status,
        "next_actions": [
            "采购经理补齐缺价格行，尤其是蔬菜菌菇和预包装类缺口。",
            "厨师长优先确认高分候选的 SOP、克重、损耗和出餐节拍。",
            "系统将公开摘要接入 v0.8a 页面，完整原始行继续留在本地私有文件。",
        ],
    }


def write_public_js(summary: dict[str, Any], output: Path) -> None:
    output.write_text(
        "window.__STAR_CHEF_V08A_IMPORT_SUMMARY__ = "
        + json.dumps(summary, ensure_ascii=False, indent=2)
        + ";\n",
        encoding="utf-8",
    )


def write_private_json(rows: list[dict[str, Any]], meta: dict[str, Any], output: Path) -> None:
    payload = {
        "privacy_mode": "private_raw_exact_prices",
        "source_file": meta["source_file"],
        "source_sheet": meta["sheet_name"],
        "generated_at": meta["generated_at"],
        "rows": rows,
    }
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import Star Chef v0.8a supplier quote workbook.")
    parser.add_argument("--source", default=DEFAULT_SOURCE, help="Path to supplier quote workbook.")
    parser.add_argument("--sheet", default="上海", help="Worksheet name to import.")
    parser.add_argument("--public-js", default="v08a-supplier-import-summary.js", help="Sanitized public JS output.")
    parser.add_argument("--private-json", default="data/private/v08a-supplier-import-raw.json", help="Private raw JSON output.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    source = Path(args.source)
    public_js = Path(args.public_js)
    private_json = Path(args.private_json)

    rows, meta = read_sheet_rows(source, args.sheet)
    summary = build_summary(rows, meta)
    write_public_js(summary, public_js)
    write_private_json(rows, meta, private_json)

    print(json.dumps({
        "public_js": str(public_js),
        "private_json": str(private_json),
        "row_count": summary["summary"]["row_count"],
        "missing_price_count": summary["summary"]["missing_price_count"],
        "missing_spec_count": summary["summary"]["missing_spec_count"],
        "public_candidate_count": len(summary["public_candidates"]),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
